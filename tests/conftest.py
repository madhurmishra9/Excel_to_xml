"""Shared fixtures and sample data for the test suite."""

import os
import sys

import pytest

# Make the package importable when tests are run from the repo root.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a small multi-sheet .xlsx and return its path.

    Includes a sheet name with a space and numeric/date cells to exercise
    sheet-name sanitisation and value stringification.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Data")
    ws1.append(["name", "qty", "price"])
    ws1.append(["apple", 3, 1.5])
    ws1.append(["banana", 12, 0.25])

    ws2 = wb.create_sheet("My Sheet")  # name with a space
    ws2.append(["a", "b"])
    ws2.append(["x", ""])  # empty cell

    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


SAMPLE_RSS = b"""<?xml version="1.0"?>
<rss version="2.0">
  <channel>
    <title>Example Feed</title>
    <item><title>First</title><link>http://e/1</link></item>
    <item><title>Second</title><link>http://e/2</link></item>
  </channel>
</rss>"""

SAMPLE_XML = b"""<?xml version="1.0"?>
<catalog>
  <book id="1"><author>A</author><title>T1</title></book>
  <book id="2"><author>B</author><title>T2</title></book>
</catalog>"""

SAMPLE_HTML = b"""<html><body>
<table>
  <tr><th>City</th><th>Pop</th></tr>
  <tr><td>Paris</td><td>2M</td></tr>
  <tr><td>Lyon</td><td>0.5M</td></tr>
</table>
</body></html>"""
