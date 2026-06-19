"""Read and write Excel workbooks as the canonical model.

``.xlsx`` / ``.xlsm`` are handled by openpyxl (read and write); legacy ``.xls``
is read with xlrd.  openpyxl cannot read ``.xls`` and xlrd cannot write at all,
which is exactly why both libraries are used.
"""

from __future__ import annotations

import os
from typing import List

from .workbook import (
    DATE_FORMAT,
    new_model,
    safe_sheet_name,
    stringify_value,
)


def _read_xls(path: str):
    """Read a legacy ``.xls`` file using xlrd."""
    import xlrd  # imported lazily so xlsx-only users need not install it

    model = new_model()
    wb = xlrd.open_workbook(path)
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        rows: List[List[str]] = []
        for row in sheet.get_rows():
            cells: List[str] = []
            for cell in row:
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode)
                    cells.append(dt.strftime(DATE_FORMAT))
                else:
                    cells.append(stringify_value(cell.value))
            rows.append(cells)
        model[sheet_name] = rows
    return model


def _read_xlsx(path: str):
    """Read a modern ``.xlsx`` / ``.xlsm`` file using openpyxl."""
    import openpyxl

    model = new_model()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([stringify_value(v) for v in row])
            model[ws.title] = rows
    finally:
        wb.close()
    return model


def read_excel(path: str):
    """Read an Excel file into the canonical model, dispatching on extension."""
    if not os.path.isfile(path):
        raise FileNotFoundError("Excel file not found: %s" % path)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _read_xls(path)
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    raise ValueError(
        "Unsupported Excel extension %r (expected .xls, .xlsx or .xlsm)" % ext
    )


def write_excel(model, path: str) -> str:
    """Write a model to a single ``.xlsx`` workbook (one sheet per model key).

    All values are written as text, per the project's "keep everything as text"
    decision.  Returns the output path.
    """
    import openpyxl

    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.isdir(parent):
        os.makedirs(parent, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove the default sheet openpyxl creates; we add our own.
    wb.remove(wb.active)

    used_names: dict = {}
    if not model:
        # An empty workbook still needs at least one sheet to be valid.
        wb.create_sheet(title="Sheet1")
    for sheet_name, rows in model.items():
        title = safe_sheet_name(str(sheet_name), used_names)
        ws = wb.create_sheet(title=title)
        for row in rows:
            ws.append(["" if cell is None else str(cell) for cell in row])

    wb.save(path)
    return path
